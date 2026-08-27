"""
Report & Analytics Service — Dashboard metrics, 30-day trends, Muster Roll matrix, 10 Report Types, and CSV Export.
"""
from __future__ import annotations
import csv
import io
from datetime import date, datetime, timedelta
from typing import List, Optional, Dict, Any, Tuple
from calendar import monthrange

from fastapi import HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_

from app.models.attendance import Attendance, AttendanceStatusEnum, AttendanceDeviceLog
from app.models.employee import Employee
from app.models.company import Department, Shift
from app.models.leave import Leave, LeaveBalance, LeaveType, LeaveStatusEnum, Holiday
from app.models.audit import AuditLog
from app.schemas.reports import (
    DashboardAnalyticsResponse, DailyTrendItem, DepartmentDistributionItem,
    ActivityFeedItem, MusterRollReportResponse, MusterRollRow, GenericReportResponse,
)
from app.utils.timezone import now_ist, today_ist


class ReportService:
    def __init__(self, db: Session):
        self.db = db

    # ------------------------------------------------------------------
    # Real-Time Dashboard Analytics
    # ------------------------------------------------------------------

    def get_dashboard_analytics(self, company_id: int) -> DashboardAnalyticsResponse:
        today = today_ist()

        # 1. Total active employees
        total_employees = (
            self.db.query(Employee)
            .filter(Employee.company_id == company_id, Employee.is_active == True)  # noqa
            .count()
        )

        # Today's attendance status counts
        counts = (
            self.db.query(Attendance.status, func.count(Attendance.id))
            .join(Employee, Attendance.employee_id == Employee.id)
            .filter(Employee.company_id == company_id, Attendance.date == today)
            .group_by(Attendance.status)
            .all()
        )

        status_map = {st: count for st, count in counts}

        present_count = status_map.get(AttendanceStatusEnum.PRESENT, 0)
        late_count = status_map.get(AttendanceStatusEnum.LATE, 0)
        half_day_count = status_map.get(AttendanceStatusEnum.HALF_DAY, 0)
        wfh_count = status_map.get(AttendanceStatusEnum.WFH, 0)
        on_duty_count = status_map.get(AttendanceStatusEnum.ON_DUTY, 0)
        on_leave_count = status_map.get(AttendanceStatusEnum.ON_LEAVE, 0)

        recorded_total = sum(status_map.values())
        absent_count = status_map.get(AttendanceStatusEnum.ABSENT, 0) + max(0, total_employees - recorded_total)

        kpis = {
            "total_employees": total_employees,
            "present_today": present_count + late_count + half_day_count,
            "absent_today": absent_count,
            "late_today": late_count,
            "on_leave_today": on_leave_count,
            "wfh_today": wfh_count,
            "on_duty_today": on_duty_count,
        }

        # 2. 30-Day Attendance Trend Line Chart
        start_date = today - timedelta(days=29)
        trend_items = []
        curr_d = start_date
        while curr_d <= today:
            d_counts = (
                self.db.query(Attendance.status, func.count(Attendance.id))
                .join(Employee, Attendance.employee_id == Employee.id)
                .filter(Employee.company_id == company_id, Attendance.date == curr_d)
                .group_by(Attendance.status)
                .all()
            )
            d_map = {st: count for st, count in d_counts}

            p = d_map.get(AttendanceStatusEnum.PRESENT, 0) + d_map.get(AttendanceStatusEnum.LATE, 0)
            a = d_map.get(AttendanceStatusEnum.ABSENT, 0)
            lt = d_map.get(AttendanceStatusEnum.LATE, 0)
            ol = d_map.get(AttendanceStatusEnum.ON_LEAVE, 0)
            wf = d_map.get(AttendanceStatusEnum.WFH, 0)

            trend_items.append(
                DailyTrendItem(
                    date=curr_d.strftime("%Y-%m-%d"),
                    present=p,
                    absent=a,
                    late=lt,
                    on_leave=ol,
                    wfh=wf,
                )
            )
            curr_d += timedelta(days=1)

        # 3. Department Distribution Breakdowns
        depts = (
            self.db.query(Department)
            .filter(Department.company_id == company_id, Department.is_active == True)  # noqa
            .all()
        )
        dept_dist = []
        for d in depts:
            emp_cnt = self.db.query(Employee).filter(Employee.department_id == d.id, Employee.is_active == True).count()  # noqa
            pres_cnt = (
                self.db.query(Attendance)
                .join(Employee, Attendance.employee_id == Employee.id)
                .filter(
                    Employee.department_id == d.id,
                    Attendance.date == today,
                    Attendance.status.in_([AttendanceStatusEnum.PRESENT, AttendanceStatusEnum.LATE, AttendanceStatusEnum.WFH]),
                )
                .count()
            )
            dept_dist.append(
                DepartmentDistributionItem(
                    department_id=d.id,
                    department_name=d.name,
                    employee_count=emp_cnt,
                    present_today=pres_cnt,
                )
            )

        # 4. Live Activity Feed (latest 10 logs)
        device_logs = (
            self.db.query(AttendanceDeviceLog)
            .join(Employee, AttendanceDeviceLog.employee_id == Employee.id)
            .filter(Employee.company_id == company_id)
            .order_by(AttendanceDeviceLog.captured_at.desc())
            .limit(10)
            .all()
        )
        activity_items = []
        for log in device_logs:
            emp = self.db.get(Employee, log.employee_id)
            action_title = "Check-In" if log.action_type.value == "check_in" else "Check-Out"
            activity_items.append(
                ActivityFeedItem(
                    id=str(log.id),
                    event_type=log.action_type.value,
                    title=f"{emp.full_name if emp else 'Employee'} logged {action_title}",
                    subtitle=f"Via {log.device_model or 'Mobile App'} · {log.ip_address or 'Local'}",
                    timestamp=log.captured_at,
                    employee_name=emp.full_name if emp else "Employee",
                    employee_code=emp.employee_code if emp else "",
                )
            )

        return DashboardAnalyticsResponse(
            kpis=kpis,
            trend_30_days=trend_items,
            department_distribution=dept_dist,
            recent_activity=activity_items,
        )

    # ------------------------------------------------------------------
    # Report #2: Monthly Muster Roll Matrix (Days 1-31 vs Employees)
    # ------------------------------------------------------------------

    def get_muster_roll(
        self,
        company_id: int,
        year: int,
        month: int,
        department_id: Optional[int] = None,
    ) -> MusterRollReportResponse:
        _, total_days = monthrange(year, month)
        from_d = date(year, month, 1)
        to_d = date(year, month, total_days)

        emp_q = self.db.query(Employee).filter(Employee.company_id == company_id, Employee.is_active == True)  # noqa
        if department_id:
            emp_q = emp_q.filter(Employee.department_id == department_id)
        employees = emp_q.order_by(Employee.employee_code).all()

        # Days with no attendance row aren't necessarily "Absent" — they may be
        # a company holiday, a weekly off, or before the employee even joined.
        # Marking every such day "A" (the old behavior) makes a brand-new
        # month, or any month with weekends, look like mass absenteeism —
        # exactly what a payroll muster roll must not do.
        holiday_days = {
            h.date.day
            for h in self.db.query(Holiday)
            .filter(Holiday.company_id == company_id, Holiday.date >= from_d, Holiday.date <= to_d)
            .all()
        }
        status_code_map = {
            "present": "P",
            "late": "L",
            "half_day": "HD",
            "wfh": "WFH",
            "on_duty": "OD",
            "on_leave": "OL",
            "holiday": "HO",
            "weekly_off": "WO",
        }

        rows = []
        for emp in employees:
            atts = (
                self.db.query(Attendance)
                .filter(
                    Attendance.employee_id == emp.id,
                    Attendance.date >= from_d,
                    Attendance.date <= to_d,
                )
                .all()
            )
            att_map = {a.date.day: a.status.value for a in atts}

            day_status_map = {}
            tot_p = 0.0
            tot_a = 0.0
            tot_l = 0.0

            for day_num in range(1, total_days + 1):
                current_date = date(year, month, day_num)
                if current_date < emp.date_of_joining:
                    code = "—"  # not yet an employee — excluded from totals
                elif day_num in att_map:
                    code = status_code_map.get(att_map[day_num], "A")
                elif day_num in holiday_days:
                    code = "HO"
                elif current_date.weekday() >= 5:  # Sat/Sun — matches the
                    code = "WO"                     # weekend convention used
                else:                                # elsewhere (seed data, Excel export)
                    code = "A"
                day_status_map[day_num] = code

                if code in ("P", "L", "WFH", "OD"):
                    tot_p += 1.0
                elif code == "HD":
                    tot_p += 0.5
                    tot_a += 0.5
                elif code in ("OL", "LWP"):
                    tot_l += 1.0
                elif code == "A":
                    tot_a += 1.0

            rows.append(
                MusterRollRow(
                    employee_id=emp.id,
                    employee_code=emp.employee_code,
                    full_name=emp.full_name,
                    department_name=emp.department.name if emp.department else None,
                    days=day_status_map,
                    total_present=tot_p,
                    total_absent=tot_a,
                    total_leave=tot_l,
                )
            )

        return MusterRollReportResponse(
            year=year,
            month=month,
            total_days=total_days,
            rows=rows,
        )

    # ------------------------------------------------------------------
    # 10 Reports Generator
    # ------------------------------------------------------------------

    def generate_report(
        self,
        report_type: str,
        company_id: int,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        department_id: Optional[int] = None,
    ) -> GenericReportResponse:
        today = today_ist()
        if not from_date:
            from_date = today.replace(day=1)
        if not to_date:
            to_date = today

        title = f"{report_type.replace('_', ' ').title()} Report"
        headers: List[str] = []
        rows: List[List[Any]] = []

        if report_type == "daily_summary":
            headers = ["Date", "Employee Code", "Name", "Department", "Check-In", "Check-Out", "Hours", "Status"]
            q = (
                self.db.query(Attendance)
                .join(Employee, Attendance.employee_id == Employee.id)
                .filter(Employee.company_id == company_id, Attendance.date >= from_date, Attendance.date <= to_date)
            )
            if department_id:
                q = q.filter(Employee.department_id == department_id)
            atts = q.order_by(Attendance.date.desc(), Employee.first_name).all()
            for a in atts:
                rows.append([
                    str(a.date),
                    a.employee.employee_code if a.employee else "",
                    a.employee.full_name if a.employee else "",
                    a.employee.department.name if a.employee and a.employee.department else "",
                    str(a.check_in_time)[:5] if a.check_in_time else "—",
                    str(a.check_out_time)[:5] if a.check_out_time else "—",
                    f"{a.working_hours}h" if a.working_hours else "—",
                    a.status.value.upper(),
                ])

        elif report_type == "late_early":
            headers = ["Date", "Employee Code", "Name", "Department", "Check-In Time", "Status", "GPS Accuracy"]
            q = (
                self.db.query(Attendance)
                .join(Employee, Attendance.employee_id == Employee.id)
                .filter(
                    Employee.company_id == company_id,
                    Attendance.date >= from_date, Attendance.date <= to_date,
                    Attendance.status.in_([AttendanceStatusEnum.LATE, AttendanceStatusEnum.HALF_DAY]),
                )
            )
            if department_id:
                q = q.filter(Employee.department_id == department_id)
            atts = q.all()
            for a in atts:
                rows.append([
                    str(a.date),
                    a.employee.employee_code if a.employee else "",
                    a.employee.full_name if a.employee else "",
                    a.employee.department.name if a.employee and a.employee.department else "",
                    str(a.check_in_time)[:5] if a.check_in_time else "—",
                    a.status.value.upper(),
                    f"{a.check_in_gps_accuracy}m" if a.check_in_gps_accuracy else "—",
                ])

        elif report_type == "leave_utilization":
            headers = ["Employee Code", "Name", "Department", "Leave Type", "Year", "Total Entitled", "Used Days", "Balance Days"]
            q = (
                self.db.query(LeaveBalance)
                .join(Employee, LeaveBalance.employee_id == Employee.id)
                .filter(Employee.company_id == company_id)
            )
            if department_id:
                q = q.filter(Employee.department_id == department_id)
            bals = q.all()
            for b in bals:
                rows.append([
                    b.employee.employee_code if b.employee else "",
                    b.employee.full_name if b.employee else "",
                    b.employee.department.name if b.employee and b.employee.department else "",
                    b.leave_type.code if b.leave_type else "",
                    b.year,
                    b.total_days,
                    b.used_days,
                    b.balance_days,
                ])

        elif report_type == "employee_master":
            headers = ["Code", "Full Name", "Email", "Phone", "Department", "Designation", "Shift", "Status"]
            q = self.db.query(Employee).filter(Employee.company_id == company_id)
            if department_id:
                q = q.filter(Employee.department_id == department_id)
            emps = q.all()
            for e in emps:
                rows.append([
                    e.employee_code,
                    e.full_name,
                    e.email,
                    e.phone or "—",
                    e.department.name if e.department else "—",
                    e.designation.name if e.designation else "—",
                    e.shift.name if e.shift else "—",
                    "Active" if e.is_active else "Inactive",
                ])

        elif report_type == "absenteeism":
            headers = ["Date", "Employee Code", "Name", "Department", "Status"]
            q = (
                self.db.query(Attendance)
                .join(Employee, Attendance.employee_id == Employee.id)
                .filter(
                    Employee.company_id == company_id,
                    Attendance.date >= from_date, Attendance.date <= to_date,
                    Attendance.status == AttendanceStatusEnum.ABSENT,
                )
            )
            if department_id:
                q = q.filter(Employee.department_id == department_id)
            atts = q.order_by(Attendance.date.desc()).all()
            for a in atts:
                rows.append([
                    str(a.date),
                    a.employee.employee_code if a.employee else "",
                    a.employee.full_name if a.employee else "",
                    a.employee.department.name if a.employee and a.employee.department else "",
                    a.status.value.upper(),
                ])

        elif report_type == "overtime":
            headers = ["Date", "Employee Code", "Name", "Department", "Shift Hours", "Worked Hours", "Overtime Hours"]
            q = (
                self.db.query(Attendance)
                .join(Employee, Attendance.employee_id == Employee.id)
                .filter(
                    Employee.company_id == company_id,
                    Attendance.date >= from_date, Attendance.date <= to_date,
                    Attendance.working_hours.isnot(None),
                )
            )
            if department_id:
                q = q.filter(Employee.department_id == department_id)
            atts = q.all()
            for a in atts:
                shift_hours = float(a.employee.shift.working_hours) if a.employee and a.employee.shift else 8.0
                worked = float(a.working_hours or 0)
                overtime = round(worked - shift_hours, 2)
                if overtime <= 0:
                    continue
                rows.append([
                    str(a.date),
                    a.employee.employee_code if a.employee else "",
                    a.employee.full_name if a.employee else "",
                    a.employee.department.name if a.employee and a.employee.department else "",
                    shift_hours,
                    worked,
                    overtime,
                ])

        elif report_type == "wfh_onduty":
            headers = ["Date", "Employee Code", "Name", "Department", "Status", "Check-In", "Check-Out"]
            q = (
                self.db.query(Attendance)
                .join(Employee, Attendance.employee_id == Employee.id)
                .filter(
                    Employee.company_id == company_id,
                    Attendance.date >= from_date, Attendance.date <= to_date,
                    Attendance.status.in_([AttendanceStatusEnum.WFH, AttendanceStatusEnum.ON_DUTY]),
                )
            )
            if department_id:
                q = q.filter(Employee.department_id == department_id)
            atts = q.order_by(Attendance.date.desc()).all()
            for a in atts:
                rows.append([
                    str(a.date),
                    a.employee.employee_code if a.employee else "",
                    a.employee.full_name if a.employee else "",
                    a.employee.department.name if a.employee and a.employee.department else "",
                    a.status.value.upper(),
                    str(a.check_in_time)[:5] if a.check_in_time else "—",
                    str(a.check_out_time)[:5] if a.check_out_time else "—",
                ])

        elif report_type == "audit_trail":
            headers = ["Timestamp", "User Email", "Action", "Entity Type", "Entity ID", "IP Address"]
            from app.models.user import User
            logs = (
                self.db.query(AuditLog)
                .join(User, AuditLog.user_id == User.id, isouter=True)
                .filter(AuditLog.created_at >= datetime.combine(from_date, datetime.min.time()),
                        AuditLog.created_at <= datetime.combine(to_date, datetime.max.time()))
                .order_by(AuditLog.created_at.desc())
                .limit(500)
                .all()
            )
            for log in logs:
                user = self.db.get(User, log.user_id) if log.user_id else None
                rows.append([
                    log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    user.email if user else "system",
                    log.action,
                    log.entity_type or "—",
                    log.entity_id if log.entity_id is not None else "—",
                    log.ip_address or "—",
                ])

        elif report_type == "shift_compliance":
            headers = ["Employee Code", "Name", "Department", "Shift", "Shift Start", "Shift End", "Grace (min)"]
            q = self.db.query(Employee).filter(Employee.company_id == company_id, Employee.is_active == True)  # noqa
            if department_id:
                q = q.filter(Employee.department_id == department_id)
            emps = q.all()
            for e in emps:
                if not e.shift:
                    continue
                rows.append([
                    e.employee_code,
                    e.full_name,
                    e.department.name if e.department else "—",
                    e.shift.name,
                    str(e.shift.start_time)[:5],
                    str(e.shift.end_time)[:5],
                    e.shift.grace_period_minutes,
                ])

        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unknown report_type '{report_type}'",
            )

        return GenericReportResponse(
            report_type=report_type,
            title=title,
            generated_at=now_ist(),
            headers=headers,
            rows=rows,
            total_records=len(rows),
        )

    def export_csv(
        self,
        report_type: str,
        company_id: int,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        department_id: Optional[int] = None,
        year: Optional[int] = None,
        month: Optional[int] = None,
    ) -> str:
        output = io.StringIO()
        writer = csv.writer(output)

        # generate_report() has no "muster_roll" branch — exporting it used to
        # 400 with "Unknown report_type 'muster_roll'" no matter what. Build
        # its CSV straight from the same grid the on-screen table uses.
        if report_type == "muster_roll":
            today = today_ist()
            data = self.get_muster_roll(
                company_id=company_id,
                year=year if year is not None else today.year,
                month=month if month is not None else today.month,
                department_id=department_id,
            )
            writer.writerow(["Emp Code", "Name"] + [str(d) for d in range(1, data.total_days + 1)] + ["Present", "Absent", "Leave"])
            for row in data.rows:
                writer.writerow([row.employee_code, row.full_name] + [row.days.get(d, "—") for d in range(1, data.total_days + 1)] + [row.total_present, row.total_absent, row.total_leave])
            return output.getvalue()

        rep = self.generate_report(report_type, company_id, from_date, to_date, department_id)
        writer.writerow(rep.headers)
        for r in rep.rows:
            writer.writerow(r)
        return output.getvalue()

    def export_excel(
        self,
        report_type: str,
        company_id: int,
        from_date: Optional[date] = None,
        to_date: Optional[date] = None,
        department_id: Optional[int] = None,
        year: Optional[int] = None,
        month: Optional[int] = None,
    ) -> bytes:
        """Generic .xlsx export — same data as export_csv/get_muster_roll, just
        styled like the per-employee Excel report instead of plain CSV."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active

        primary_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(*(Side(style="thin", color="E2E8F0") for _ in range(4)))
        center = Alignment(horizontal="center", vertical="center")

        if report_type == "muster_roll":
            today = today_ist()
            data = self.get_muster_roll(
                company_id=company_id,
                year=year if year is not None else today.year,
                month=month if month is not None else today.month,
                department_id=department_id,
            )
            ws.title = f"Muster Roll {data.month}-{data.year}"
            headers = ["Emp Code", "Name"] + [str(d) for d in range(1, data.total_days + 1)] + ["Present", "Absent", "Leave"]
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = primary_fill
                cell.alignment = center
                cell.border = thin_border
            for r_idx, row in enumerate(data.rows, start=2):
                values = [row.employee_code, row.full_name] + [row.days.get(d, "—") for d in range(1, data.total_days + 1)] + [row.total_present, row.total_absent, row.total_leave]
                for col, v in enumerate(values, start=1):
                    cell = ws.cell(row=r_idx, column=col, value=v)
                    cell.border = thin_border
                    if col > 2:
                        cell.alignment = center
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 22
            for i in range(3, 3 + data.total_days):
                ws.column_dimensions[get_column_letter(i)].width = 5
            ws.freeze_panes = "C2"
        else:
            rep = self.generate_report(report_type, company_id, from_date, to_date, department_id)
            ws.title = rep.title[:31]  # Excel sheet-name length limit
            for col, h in enumerate(rep.headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = primary_fill
                cell.alignment = center
                cell.border = thin_border
            for r_idx, row in enumerate(rep.rows, start=2):
                for col, v in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=col, value=v)
                    cell.border = thin_border
            for i, h in enumerate(rep.headers, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(h) + 4))
            ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # Per-Employee Excel Report — search an employee, pick a date range
    # (month or custom/week), export full attendance + daily task summary.
    # ------------------------------------------------------------------

    def generate_employee_excel_report(
        self,
        employee_id: int,
        from_date: date,
        to_date: date,
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        employee = self.db.get(Employee, employee_id)
        if not employee:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")

        atts = (
            self.db.query(Attendance)
            .filter(
                Attendance.employee_id == employee_id,
                Attendance.date >= from_date,
                Attendance.date <= to_date,
            )
            .order_by(Attendance.date)
            .all()
        )
        att_by_date = {a.date: a for a in atts}

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        primary_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        title_font = Font(bold=True, size=14, color="1E1B4B")
        label_font = Font(bold=True, size=10, color="475569")
        thin_border = Border(*(Side(style="thin", color="E2E8F0") for _ in range(4)))
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- Header block: employee identity + period ---
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Attendance Report — {employee.full_name} ({employee.employee_code})"
        ws["A1"].font = title_font

        ws["A2"] = "Department:"
        ws["A2"].font = label_font
        ws["B2"] = employee.department.name if employee.department else "—"
        ws["D2"] = "Designation:"
        ws["D2"].font = label_font
        ws["E2"] = employee.designation.name if employee.designation else "—"
        ws["A3"] = "Period:"
        ws["A3"].font = label_font
        ws["B3"] = f"{from_date.strftime('%d %b %Y')} to {to_date.strftime('%d %b %Y')}"

        # --- Table header row ---
        headers = ["Date", "Day", "Check-In", "Check-Out", "Working Hours", "Status", "Tasks Completed"]
        header_row = 5
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = header_font
            cell.fill = primary_fill
            cell.alignment = center
            cell.border = thin_border

        # --- Data rows ---
        totals = {"present": 0.0, "absent": 0.0, "late": 0, "leave": 0, "hours": 0.0}
        row_idx = header_row + 1
        curr = from_date
        while curr <= to_date:
            att = att_by_date.get(curr)
            st = att.status.value if att else "absent"
            if st in ("present", "late", "wfh", "on_duty"):
                totals["present"] += 1
            elif st == "half_day":
                totals["present"] += 0.5
                totals["absent"] += 0.5
            elif st in ("on_leave", "lwp"):
                totals["leave"] += 1
            elif st == "absent":
                totals["absent"] += 1
            if st == "late":
                totals["late"] += 1
            if att and att.working_hours is not None:
                totals["hours"] += float(att.working_hours)

            values = [
                curr.strftime("%d-%b-%Y"),
                curr.strftime("%A"),
                str(att.check_in_time)[:5] if att and att.check_in_time else "—",
                str(att.check_out_time)[:5] if att and att.check_out_time else "—",
                float(att.working_hours) if att and att.working_hours is not None else "—",
                st.replace("_", " ").upper(),
                (att.checkout_task_summary if att and att.checkout_task_summary else "—"),
            ]
            for col, v in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=v)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 7))
                if curr.weekday() >= 5:  # weekend tint
                    cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            row_idx += 1
            curr += timedelta(days=1)

        # --- Summary block ---
        summary_row = row_idx + 1
        ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=12)
        summary_items = [
            ("Present Days", totals["present"]),
            ("Absent Days", totals["absent"]),
            ("Late Days", totals["late"]),
            ("Leave Days", totals["leave"]),
            ("Total Working Hours", round(totals["hours"], 2)),
        ]
        for i, (label, value) in enumerate(summary_items):
            r = summary_row + 1 + i
            ws.cell(row=r, column=1, value=label).font = label_font
            ws.cell(row=r, column=2, value=value)

        # --- Column widths ---
        widths = [14, 12, 11, 11, 14, 12, 45]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = f"A{header_row + 1}"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
